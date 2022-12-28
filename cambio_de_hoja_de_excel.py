# -*- coding: utf-8 -*-
"""
Created on Wed Nov  9 16:37:15 2022

@author: Ivan Tonatiuh
"""

from openpyxl import load_workbook

df.to_excel("template.xlsx")

#%%
wb2 = load_workbook('template.xlsx')
wb2.create_sheet('2b')


wb2.save('template.xlsx')
#%%

df.to_excel('template.xlsx', sheet_name='2')


#%%

#create a excel writer object
with pd.ExcelWriter("template.xlsx") as writer:
   
    # use to_excel function and specify the sheet_name and index
    # to store the dataframe in specified sheet
    df.to_excel(writer, sheet_name="Fruits", index=False)
    df.to_excel(writer, sheet_name="Vegetables", index=False)
    df.to_excel(writer, sheet_name="Baked Items", index=False)

#%% forma de agregar más 

# create a excel writer object as shown using
# Excelwriter function
with pd.ExcelWriter("template.xlsx", mode="a", engine="openpyxl") as writer:
     
    # use to_excel function and specify the sheet_name and index to
    # store the dataframe in specified sheet
    df.to_excel(writer, sheet_name="Cool drinks")
    df.to_excel(writer, sheet_name="yes")

